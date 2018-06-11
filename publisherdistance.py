#publisherdistance.py

from openpyxl import *




def GetPublisherRecord(title):
	wb = load_workbook(title)
	pr = {}
	publishers = []
	addresses =[]
	#for i in wb.sheetnames:
	#	print i
	ws = wb.worksheets[0]
	for col in ws.iter_cols(max_col=1):
		for cell in col:
			publishers.append(cell.value)
	for col in ws.iter_cols(min_col=2, max_col=2):
		for cell in col:
			addresses.append(cell.value)
	#print publishers
	#print addresses

	pr = dict(zip(publishers, addresses))
	print pr


	
def main():
	GetPublisherRecord("../publisher.xlsx")







main()